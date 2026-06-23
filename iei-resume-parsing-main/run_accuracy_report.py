import os
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from parsers.file_extractor import parse_file
from parsers.info_extractor import parse_extracted_entities
from utils.logger import get_logger

logger = get_logger()

def main():
    logger.info("Initializing Experience Accuracy Report Generator...")
    
    project_dir = os.path.dirname(os.path.abspath(__file__))
    dataset_dir = os.path.join(project_dir, "dataset")
    
    # Initialize dataset directory if it does not exist
    if not os.path.exists(dataset_dir):
        os.makedirs(dataset_dir)
        logger.info(f"Created dataset directory at: {dataset_dir}")
        
    expected_json_path = os.path.join(dataset_dir, "expected_experience.json")
    
    # Initialize expected experience placeholders if not present
    if not os.path.exists(expected_json_path):
        sample_expected = {
            "example_cv_1.pdf": 10.0,
            "example_cv_2.docx": 5.0
        }
        with open(expected_json_path, "w") as f:
            json.dump(sample_expected, f, indent=2)
        logger.info(f"Created expected experience metadata placeholder at: {expected_json_path}")
        
    # Read expected experience mappings
    try:
        with open(expected_json_path, "r") as f:
            expected_mapping = json.load(f)
    except Exception as e:
        logger.error(f"Failed to read expected_experience.json: {e}")
        expected_mapping = {}
        
    # Scan for files in dataset
    supported_extensions = (".pdf", ".docx", ".doc", ".png", ".jpg", ".jpeg")
    files = [f for f in os.listdir(dataset_dir) if f.lower().endswith(supported_extensions)]
    
    if not files:
        print("\n" + "="*80)
        print("DATASET DIRECTORY IS EMPTY OR CONTAINS NO SUPPORTED RESUMES.")
        print(f"Please place candidate CVs inside: {dataset_dir}")
        print(f"And define their expected experience in: {expected_json_path}")
        print("Example JSON format:")
        print(json.dumps({"candidate_cv.pdf": 12}, indent=2))
        print("="*80 + "\n")
        
        # Create a blank report file to avoid empty errors
        report_path = os.path.join(project_dir, "experience_accuracy_report.xlsx")
        df_empty = pd.DataFrame(columns=["Resume", "Expected Experience", "Predicted Experience", "Difference", "Status"])
        df_empty.to_excel(report_path, index=False)
        logger.info(f"Generated empty experience_accuracy_report.xlsx at: {report_path}")
        return
        
    logger.info(f"Found {len(files)} resumes to evaluate in dataset folder.")
    
    results = []
    
    for filename in files:
        file_path = os.path.join(dataset_dir, filename)
        logger.info(f"Evaluating: {filename}...")
        
        # Retrieve expected experience
        expected_exp = expected_mapping.get(filename, 0.0)
        
        try:
            # 1. Text extraction
            payload = parse_file(file_path, logger)
            
            # 2. Hybrid Timeline parsing
            profile, _ = parse_extracted_entities(payload, logger)
            
            # Retrieve predicted total professional experience
            predicted_exp = profile.get("total_experience_years", 0.0)
            
            diff = predicted_exp - expected_exp
            status = "Match" if abs(diff) < 0.1 else "Mismatch"
            
            results.append({
                "Resume": filename,
                "Expected Experience": expected_exp,
                "Predicted Experience": predicted_exp,
                "Difference": round(diff, 1),
                "Status": status
            })
            logger.info(f"Successfully evaluated: {filename} (Expected: {expected_exp}, Predicted: {predicted_exp}, Status: {status})")
        except Exception as e:
            logger.error(f"Failed to evaluate {filename}: {e}")
            results.append({
                "Resume": filename,
                "Expected Experience": expected_exp,
                "Predicted Experience": 0.0,
                "Difference": -expected_exp,
                "Status": "Failed"
            })
            
    # Write styled Excel Workbook
    report_path = os.path.join(project_dir, "experience_accuracy_report.xlsx")
    logger.info(f"Writing styled workbook to {report_path}...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Experience Accuracy"
    
    # Style Definitions
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    match_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
    mismatch_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # soft orange
    failed_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid") # soft red
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Write Headers
    headers = ["Resume", "Expected Experience (Yrs)", "Predicted Experience (Yrs)", "Difference (Yrs)", "Status"]
    ws.append(headers)
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx > 1 else align_left
        cell.border = thin_border
        
    # Write Rows
    for row_idx, r in enumerate(results, start=2):
        row_data = [
            r["Resume"],
            r["Expected Experience"],
            r["Predicted Experience"],
            r["Difference"],
            r["Status"]
        ]
        ws.append(row_data)
        
        # Apply fonts, borders, alignments, and status fills
        status = r["Status"]
        row_fill = None
        if status == "Match":
            row_fill = match_fill
        elif status == "Mismatch":
            row_fill = mismatch_fill
        else:
            row_fill = failed_fill
            
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name=font_family, size=10)
            cell.border = thin_border
            cell.fill = row_fill
            
            # Alignments
            if col_idx == 1:
                cell.alignment = align_left
            elif col_idx in [2, 3, 4]:
                cell.alignment = align_right
            else:
                cell.alignment = align_center
                
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(report_path)
    logger.info("Experience Accuracy Report compiled successfully.")
    print(f"\nReport written to: {report_path}\n")

if __name__ == "__main__":
    main()

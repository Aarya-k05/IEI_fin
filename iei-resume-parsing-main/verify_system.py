import os
import sys

# Add project root to sys.path
sys.path.append(os.path.dirname(__file__))

from database.db_manager import init_db, save_candidate_pipeline_results, get_all_candidates_ranked, clear_all_candidates, get_candidate_details
from parsers.info_extractor import parse_extracted_entities
from scoring.scoring_engine import calculate_candidate_score
from exports.excel_generator import build_evaluation_excel
from utils.logger import get_logger

logger = get_logger()

# 1. Define sample CV Mock Strings
MOCK_CV_1 = """
DR. AMIT KUMAR SHARMA
Email: amit.sharma@email.com
Phone: +91 98765 43210
Address: Delhi, India

RESEARCH INTERESTS
Machine Learning, Natural Language Processing, Faculty Evaluation

EDUCATION
- PhD in Computer Science & Engineering, IIT Delhi, 2018 - 2021
- M.Tech in Computer Science, BITS Pilani, 2015 - 2017
- B.Tech in CSE, DTU Delhi, 2011 - 2015

TEACHING EXPERIENCE
- Assistant Professor, Department of CSE, DTU Delhi: July 2021 - Present
- Lecturer, Department of IT, DTU: July 2018 - June 2021

PUBLICATIONS
1. A. K. Sharma, "Deep Learning for Resume Analysis," IEEE Transactions on AI, 2023.
2. A. K. Sharma et al., "Evaluation Frameworks in Academia," Springer Journal of Education, 2022.
3. "Parsing Academic CVs via Heuristics," Elsevier Computer Science Review, 2021.
4. "Scopus Indexed Publication on Neural Networks," Journal of Web Engineering, 2020.
5. "SCI Indexed Paper on Automated Grading Systems," ACM Conference on SIGCSE, 2019.
6. "A review of Natural Language Processing," UGC Care List Journal, 2018.

RESEARCH GUIDANCE
- Guided 3 PhD Scholars in ML
- Supervised 8 PG students (M.Tech)
- Handled 2 research projects sponsored by DST

PATENTS
- Granted Patent: Patent No. IN384729-B (Method for automated resume scoring, 2023)
- Filed Patent: Application No. 202211028392 (Dynamic web analytics, 2022)

EXTRA CURRICULAR ACTIVITIES
- Active NSS Volunteer during college days
- Event Coordinator for DTU Cultural Fest, 2022
- Hackathon Mentor and Coordinator
- Attended FDP on Deep Learning, IIT Bombay, 2023
"""

MOCK_CV_2 = """
PRIYA PATEL
Email: priya.patel@webmail.edu
Phone: 8765432109

OBJECTIVE
Motivated software engineer looking for academic faculty positions.

EDUCATION
- M.Tech in Information Technology, NIT Trichy, 2019
- B.E. in Information Technology, Pune University, 2017

TEACHING EXPERIENCE
- Lecturer, NIT Trichy: 2019 - 2022

INDUSTRY EXPERIENCE
- Software Engineer, Tata Consultancy Services: June 2017 - June 2019

PUBLICATIONS
1. P. Patel, "Blockchain and IoT Security," Springer Journal, 2021.
2. P. Patel, "A Survey of Smart Contracts," IEEE Conference on Cyber Security, 2020.

RESEARCH GUIDANCE
- Guided 2 PG students
- Completed 1 funded project

PATENTS
- Filed Patent: Application No. 20214104938 (Smart contracts framework)

CERTIFICATIONS
- Attended STTP on Data Science, NIT Trichy, 2020

EXTRA CURRICULAR ACTIVITIES
- Event Coordinator, NIT Trichy Fest
- Member of IEEE Student Branch
"""

MOCK_CV_KAVITHA = """
D.Kavitha
Sr.Asst professor
Department of Information Technology
PVP Siddhartha Institute of Technology
Kanuru.

Academics:
Ph.D(CSE) | JNTU Campus College | JNTU, Kakinada | December,2018 | 
M.Tech(CSE) | Acharya Nagarjuna University | Acharya Nagarjuna University | 2003-2005 | Distinction with 82%
MCA | KBN College | IGNOU | 2002 | 1st Class with 64.89 %
BCA | KBN College | IGNOU | 2002 | 1st Class with 61.92%

Experience:
- Working as Sr.Assistant professor from March 2010 in Prasad V. Potluri Siddhartha Institute of Technology, Vijayawada.
- Worked as Assistant professor from Dec 3rd, 2004to Feb 2010 in Prasad V. Potluri Siddhartha Institute of Technology, Vijayawada.
- Ratified as an Assistant professor from JNTU Kakinada on April 2010
"""

MOCK_CV_OVERLAP = """
DR. VIPIN RAJ
Email: vipin.raj@iitb.ac.in

EDUCATION
- PhD in Computer Science, IIT Bombay, 2020

TEACHING EXPERIENCE
- Professor, Department of CSE, IIT Bombay: 2021 - Present
- Joint Professor, Centre for AI, IIT Bombay: 2022 - 2024
"""

MOCK_CV_TIMELINE_REDESIGN = """
EXPERIENCE REDESIGN TEST PROFILE
TEACHING EXPERIENCE
- Assistant Professor, IIT Jammu, February 01, 2018 - till date
- Scientist D, July 01, 2014 - January 31, 2018
- Scientist C, July 01, 2010 - June 30, 2014
- Postdoctoral Fellow, September 08, 2011 - December 31, 2012
- Scientist B, February 28, 2007 - June 30, 2010
- Teaching Assistant, July 2003 - February 2007
"""

def run_verification():
    logger.info("Initializing Verification Run...")
    
    # Patch extract_candidate_profile_via_llm to mock Ollama behavior offline.
    # This now drives BOTH personal info (name/email/phone) and employment
    # records from a single mocked call, matching the real holistic pipeline.
    import parsers.info_extractor
    original_llm_extractor = parsers.info_extractor.extract_candidate_profile_via_llm
    
    def mock_llm_extractor(text, model="qwen2.5:3b"):
        text_lower = text.lower()
        if "vipin" in text_lower:
            return {
                "name": None,   # left unset on purpose: exercises the layout_analysis fallback path
                "email": None,  # left unset on purpose: exercises the regex select_best_email fallback path
                "phone": None,
                "employment_records": [
                    {
                        "designation": "Professor",
                        "organization": "IIT Bombay",
                        "start_date": "2021-06",
                        "end_date": "Present",
                        "employment_type": "Academic"
                    },
                    {
                        "designation": "Professor",
                        "organization": "IIT Bombay",
                        "start_date": "2022-06",
                        "end_date": "2024-06",
                        "employment_type": "Academic"
                    }
                ]
            }
        elif "experience redesign test profile" in text_lower:
            return {
                "name": None,
                "email": None,
                "phone": None,
                "employment_records": [
                    {
                        "designation": "Assistant Professor",
                        "organization": "IIT Jammu",
                        "start_date": "February 01, 2018",
                        "end_date": "till date",
                        "employment_type": "Academic"
                    },
                    {
                        "designation": "Scientist D",
                        "organization": "Unknown",
                        "start_date": "July 01, 2014",
                        "end_date": "January 31, 2018",
                        "employment_type": "Research"
                    },
                    {
                        "designation": "Scientist C",
                        "organization": "Unknown",
                        "start_date": "July 01, 2010",
                        "end_date": "June 30, 2014",
                        "employment_type": "Research"
                    },
                    {
                        "designation": "Postdoctoral Fellow",
                        "organization": "Unknown",
                        "start_date": "September 08, 2011",
                        "end_date": "December 31, 2012",
                        "employment_type": "Research"
                    },
                    {
                        "designation": "Scientist B",
                        "organization": "Unknown",
                        "start_date": "July 01, 2007",
                        "end_date": "June 30, 2010",
                        "employment_type": "Research"
                    },
                    {
                        "designation": "Teaching Assistant",
                        "organization": "Unknown",
                        "start_date": "July 2003",
                        "end_date": "July 01, 2007",
                        "employment_type": "Academic"
                    }
                ]
            }
        elif "amit" in text_lower:
            # This branch exercises the NEW holistic path: name/email/phone come
            # from the (mocked) LLM call instead of the regex fallback chain.
            return {
                "name": "Amit Kumar Sharma",
                "email": "amit.sharma@email.com",
                "phone": "+91 98765 43210",
                "employment_records": [
                    {
                        "designation": "Assistant Professor",
                        "organization": "DTU Delhi",
                        "start_date": "July 2021",
                        "end_date": "Present",
                        "employment_type": "Academic"
                    },
                    {
                        "designation": "Lecturer",
                        "organization": "DTU",
                        "start_date": "July 2018",
                        "end_date": "June 2021",
                        "employment_type": "Academic"
                    }
                ],
                "education_records": [
                    {"level": "PhD", "degree": "PhD", "institution": "IIT Delhi", "year": "2021"},
                    {"level": "PG", "degree": "M.Tech", "institution": "BITS Pilani", "year": "2017"},
                    {"level": "UG", "degree": "B.Tech", "institution": "DTU Delhi", "year": "2015"},
                ],
                "publication_records": [
                    {"title": "Deep Learning for Resume Analysis", "journal_name": "IEEE Transactions on AI",
                     "published_under": "IEEE", "year": "2023", "impact_factor": None, "scopus_indexed": None},
                    {"title": "Evaluation Frameworks in Academia", "journal_name": "Springer Journal of Education",
                     "published_under": "Springer", "year": "2022", "impact_factor": None, "scopus_indexed": None},
                    # Intentionally fabricated entry the mock LLM "hallucinates" - must be rejected by grounding
                    {"title": "Quantum Computing Approaches to Faculty Hiring", "journal_name": "Made Up Quarterly",
                     "published_under": None, "year": "2099", "impact_factor": "50.0", "scopus_indexed": "Yes"},
                ]
            }
        elif "priya" in text_lower:
            return {
                "name": "Priya Patel",
                "email": "priya.patel@webmail.edu",
                "phone": "8765432109",
                "employment_records": [
                    {
                        "designation": "Lecturer",
                        "organization": "NIT Trichy",
                        "start_date": "2019",
                        "end_date": "2022",
                        "employment_type": "Academic"
                    },
                    {
                        "designation": "Software Engineer",
                        "organization": "Tata Consultancy Services",
                        "start_date": "June 2017",
                        "end_date": "June 2019",
                        "employment_type": "Industry"
                    }
                ],
                "education_records": [
                    {"level": "PG", "degree": "M.Tech", "institution": "NIT Trichy", "year": "2019"},
                    {"level": "UG", "degree": "B.E.", "institution": "Pune University", "year": "2017"},
                ],
                "publication_records": [
                    {"title": "Blockchain and IoT Security", "journal_name": "Springer Journal",
                     "published_under": "Springer", "year": "2021", "impact_factor": None, "scopus_indexed": None},
                    {"title": "A Survey of Smart Contracts", "journal_name": "IEEE Conference on Cyber Security",
                     "published_under": "IEEE", "year": "2020", "impact_factor": None, "scopus_indexed": None},
                ]
            }
        elif "kavitha" in text_lower:
            return {
                "name": None,   # left unset on purpose: exercises layout_analysis fallback (potential_name="D.Kavitha")
                "email": None,
                "phone": None,
                "employment_records": [
                    {
                        "designation": "Sr Assistant Professor",
                        "organization": "Prasad V. Potluri Siddhartha Institute of Technology",
                        "start_date": "March 2010",
                        "end_date": "Present",
                        "employment_type": "Academic"
                    },
                    {
                        "designation": "Assistant Professor",
                        "organization": "Prasad V. Potluri Siddhartha Institute of Technology",
                        "start_date": "Dec 3rd, 2004",
                        "end_date": "Feb 2010",
                        "employment_type": "Academic"
                    }
                ]
            }
        return None
        
    parsers.info_extractor.extract_candidate_profile_via_llm = mock_llm_extractor
    
    # Init DB
    init_db()
    
    # Clear any previous candidates to start clean
    clear_all_candidates()
    logger.info("Database cleared.")
    
    # Process Mock CV 1
    logger.info("Processing Mock Candidate 1 (Dr. Amit Kumar Sharma)...")
    payload1 = {
        "raw_text": MOCK_CV_1,
        "potential_name": "Dr. Amit Kumar Sharma",
        "file_type": "PDF",
        "filename": "amit_sharma_cv.pdf"
    }
    profile1, sections1 = parse_extracted_entities(payload1, logger)
    scores1 = calculate_candidate_score(profile1)
    
    meta1 = {
        "filename": payload1["filename"],
        "file_type": payload1["file_type"],
        "raw_text": payload1["raw_text"],
        "sections": sections1
    }
    cid1 = save_candidate_pipeline_results(profile1, meta1, scores1)
    logger.info(f"Saved Candidate 1 to DB. ID: {cid1}")
    
    # Process Mock CV 2
    logger.info("Processing Mock Candidate 2 (Priya Patel)...")
    payload2 = {
        "raw_text": MOCK_CV_2,
        "potential_name": "Priya Patel",
        "file_type": "PDF",
        "filename": "priya_patel_cv.pdf"
    }
    profile2, sections2 = parse_extracted_entities(payload2, logger)
    scores2 = calculate_candidate_score(profile2)
    
    meta2 = {
        "filename": payload2["filename"],
        "file_type": payload2["file_type"],
        "raw_text": payload2["raw_text"],
        "sections": sections2
    }
    cid2 = save_candidate_pipeline_results(profile2, meta2, scores2)
    logger.info(f"Saved Candidate 2 to DB. ID: {cid2}")
    
    # Process Mock CV 3 (Dr. Kavitha)
    logger.info("Processing Mock Candidate 3 (D.Kavitha)...")
    payload3 = {
        "raw_text": MOCK_CV_KAVITHA,
        "potential_name": "D.Kavitha",
        "file_type": "PDF",
        "filename": "kavitha_cv.pdf"
    }
    profile3, sections3 = parse_extracted_entities(payload3, logger)
    scores3 = calculate_candidate_score(profile3)
    
    meta3 = {
        "filename": payload3["filename"],
        "file_type": payload3["file_type"],
        "raw_text": payload3["raw_text"],
        "sections": sections3
    }
    cid3 = save_candidate_pipeline_results(profile3, meta3, scores3)
    logger.info(f"Saved Candidate 3 to DB. ID: {cid3}")
    
    # Process Mock CV 4 (Vipin Raj - Overlap case)
    logger.info("Processing Mock Candidate 4 (Dr. Vipin Raj - Overlapping Experience)...")
    payload4 = {
        "raw_text": MOCK_CV_OVERLAP,
        "potential_name": "Dr. Vipin Raj",
        "file_type": "PDF",
        "filename": "vipin_raj_cv.pdf"
    }
    profile4, sections4 = parse_extracted_entities(payload4, logger)
    scores4 = calculate_candidate_score(profile4)
    
    meta4 = {
        "filename": payload4["filename"],
        "file_type": payload4["file_type"],
        "raw_text": payload4["raw_text"],
        "sections": sections4
    }
    cid4 = save_candidate_pipeline_results(profile4, meta4, scores4)
    logger.info(f"Saved Candidate 4 to DB. ID: {cid4}")
    
    # Process Mock CV 5 (Timeline Redesign case)
    logger.info("Processing Mock Candidate 5 (Experience Redesign Test Profile)...")
    payload5 = {
        "raw_text": MOCK_CV_TIMELINE_REDESIGN,
        "potential_name": "Test Profile",
        "file_type": "PDF",
        "filename": "redesign_test_cv.pdf"
    }
    profile5, sections5 = parse_extracted_entities(payload5, logger)
    scores5 = calculate_candidate_score(profile5)
    
    meta5 = {
        "filename": payload5["filename"],
        "file_type": payload5["file_type"],
        "raw_text": payload5["raw_text"],
        "sections": sections5
    }
    cid5 = save_candidate_pipeline_results(profile5, meta5, scores5)
    logger.info(f"Saved Candidate 5 to DB. ID: {cid5}")
    
    # Retrieve Ranked Candidates
    ranked_candidates = get_all_candidates_ranked()
    
    logger.info(f"Ranked candidates retrieved: {len(ranked_candidates)}")
    for cand in ranked_candidates:
        print(f"Rank {cand['rank']}: {cand['name']} | Degree: {cand['highest_qualification']} | Exp (Teaching): {cand['teaching_experience_years']} yrs | Total Exp: {cand['total_experience_years']} yrs | Score: {cand['total_score']}")
        
    # Build Excel
    excel_out_path = os.path.join(os.path.dirname(__file__), "ranked_candidates.xlsx")
    logger.info(f"Generating Excel workbook at: {excel_out_path}")
    build_evaluation_excel(ranked_candidates, excel_out_path)
    logger.info("Excel compilation complete.")
    
    # Verification Assertions
    assert len(ranked_candidates) == 5, f"Should have 5 candidates in DB, got {len(ranked_candidates)}"
    
    # Check D.Kavitha details (Name should be cleaned/normalized with spaces: "D. Kavitha")
    kavitha = next(c for c in ranked_candidates if c["name"] == "D. Kavitha")
    kav_details = get_candidate_details(kavitha["id"])
    profile3_db = kav_details["extracted_info"]
    
    assert kavitha["highest_qualification"] == "PhD", "Kavitha highest qualification should be PhD"
    assert kavitha["teaching_experience_years"] == 21.0, f"Kavitha teaching experience years expected 21.0, got {kavitha['teaching_experience_years']}"
    
    # Verify structured positions timeline for D. Kavitha
    positions = profile3_db.get("positions", [])
    assert len(positions) == 2, f"Expected 2 positions, got {len(positions)}"
    
    # Assert Position 1
    assert positions[0]["designation"] == "Assistant Professor", f"Expected Assistant Professor, got {positions[0]['designation']}"
    assert positions[0]["start_date"] == "2004-12-03", f"Expected 2004-12-03, got {positions[0]['start_date']}"
    assert positions[0]["end_date"] == "2010-02-01", f"Expected 2010-02-01, got {positions[0]['end_date']}"
    
    # Check Overlap Candidate (Vipin Raj) to ensure interval union logic worked without double counting
    vipin = next(c for c in ranked_candidates if c["name"] == "Vipin Raj")
    assert vipin["teaching_experience_years"] == 5.0, f"Expected Vipin Raj to have 5.0 years experience, got {vipin['teaching_experience_years']}"
    
    # Check Redesign Test Candidate (Candidate 5)
    redesign_cand = next(c for c in ranked_candidates if c["name"] == "Test Profile")
    assert redesign_cand["total_experience_years"] == 22, f"Expected Test Profile to have 22 years total exp, got {redesign_cand['total_experience_years']}"
    assert redesign_cand["teaching_experience_years"] == 12, f"Expected Test Profile to have 12 years academic exp, got {redesign_cand['teaching_experience_years']}"
    assert redesign_cand["research_experience_years"] == 10, f"Expected Test Profile to have 10 years research exp, got {redesign_cand['research_experience_years']}"
    
    # Check that Explanation Report contains "EXTRACTION CONFIDENCE SCORECARD"
    vipin_details = get_candidate_details(vipin["id"])
    report_text = vipin_details["explanation_report"]
    assert "EXTRACTION CONFIDENCE SCORECARD" in report_text, "Explanation report must contain Extraction Confidence Scorecard"
    assert "Candidate Name Extraction:" in report_text, "Explanation report must detail Name Extraction confidence"

    # Check Amit Kumar Sharma to confirm the NEW holistic LLM path (name/email/phone
    # extracted in the same call as employment records) is actually wired up and
    # the result is grounded against the raw resume text, not just trusted blindly.
    amit = next(c for c in ranked_candidates if c["name"] == "Amit Kumar Sharma")
    amit_details = get_candidate_details(amit["id"])
    profile1_db = amit_details["extracted_info"]
    assert profile1_db["email"] == "amit.sharma@email.com", f"Expected LLM-grounded email, got {profile1_db['email']}"
    assert profile1_db["phone"] == "+91 98765 43210", f"Expected LLM-grounded phone, got {profile1_db['phone']}"
    assert profile1_db["confidence_scores"]["name"]["score"] == 97, "Amit's name should come from the llm_holistic path (score 97)"

    # New: education institutions per degree level, resolved + grounded
    assert profile1_db["education"]["ug_institution"] == "DTU Delhi", f"Expected UG institution DTU Delhi, got {profile1_db['education']}"
    assert profile1_db["education"]["pg_institution"] == "BITS Pilani", f"Expected PG institution BITS Pilani, got {profile1_db['education']}"
    assert profile1_db["education"]["phd_institution"] == "IIT Delhi", f"Expected PhD institution IIT Delhi, got {profile1_db['education']}"

    # New: industry experience has its own bucket now, separate from teaching/research/admin
    assert profile1_db["industry_experience_years"] == 0.0, "Amit has no industry roles in his mock CV - should be 0"

    # New: FDP/STTP count replaces the old "Extra-Curricular" marking-sheet column
    assert profile1_db["fdp_sttp"]["fdp_sttp_count"] >= 1, f"Expected at least 1 FDP/STTP entry, got {profile1_db['fdp_sttp']}"

    # New: publication detail records - real ones kept, fabricated one rejected by grounding
    pub_titles_verified = {p["title"]: p["verified"] for p in profile1_db["publications_detail"]}
    assert pub_titles_verified.get("Deep Learning for Resume Analysis") is True, "Real publication should be verified"
    assert any(not v for t, v in pub_titles_verified.items() if "Quantum Computing" in t), \
        "Fabricated publication ('Quantum Computing Approaches to Faculty Hiring') must be flagged unverified, not trusted"

    # Check Vipin Raj still correctly falls back to layout_analysis + regex email
    # selection when the (mocked) LLM doesn't return name/email for this candidate -
    # proving the fallback chain still works when Ollama has nothing useful to say.
    profile4_db = vipin_details["extracted_info"]
    assert profile4_db["email"] == "vipin.raj@iitb.ac.in", f"Expected regex-selected email, got {profile4_db['email']}"
    assert profile4_db["confidence_scores"]["name"]["score"] == 95, "Vipin's name should come from the layout_analysis fallback (score 95)"

    # Check Priya Patel: industry experience bucket should be non-zero since her
    # mock CV has a genuine Industry-classified role (Software Engineer, TCS).
    priya = next(c for c in ranked_candidates if c["name"] == "Priya Patel")
    priya_details = get_candidate_details(priya["id"])
    profile2_db = priya_details["extracted_info"]
    assert profile2_db["industry_experience_years"] == 1.0, f"Expected ~1 year industry experience for Priya (int-truncated), got {profile2_db['industry_experience_years']}"
    assert profile2_db["fdp_sttp"]["fdp_sttp_count"] >= 1, f"Expected at least 1 FDP/STTP entry for Priya, got {profile2_db['fdp_sttp']}"
    assert profile2_db["education"]["pg_institution"] == "NIT Trichy", f"Expected PG institution NIT Trichy, got {profile2_db['education']}"

    # --- NEW: USSC Marking Scheme & Publications Detail workbooks ---
    # Build both new exports end-to-end and sanity check their actual cell content,
    # not just "didn't crash".
    from exports.excel_generator import build_marking_scheme_excel, build_publications_workbook
    import openpyxl as _openpyxl

    marking_scheme_path = os.path.join(os.path.dirname(__file__), "_test_marking_scheme.xlsx")
    publications_path = os.path.join(os.path.dirname(__file__), "_test_publications.xlsx")

    build_marking_scheme_excel(ranked_candidates, marking_scheme_path, subject="Computer Engg.", post="Professor", interview_date="26-08-2022")
    build_publications_workbook(ranked_candidates, publications_path)

    ms_wb = _openpyxl.load_workbook(marking_scheme_path)
    ms_ws = ms_wb.active
    assert ms_ws["A1"].value == "Marking Scheme at the time of USSC Interview", "Marking scheme title mismatch"
    assert "FDP / STTP" in str(ms_ws.cell(row=6, column=10).value), "FDP/STTP column header missing from marking scheme"
    # Find Amit's row and confirm his PhD institution (not a mark out of 10) is shown
    amit_row_found = any(
        ms_ws.cell(row=r, column=2).value == "Amit Kumar Sharma" and ms_ws.cell(row=r, column=5).value == "IIT Delhi"
        for r in range(7, 7 + len(ranked_candidates))
    )
    assert amit_row_found, "Expected Amit's row to show 'IIT Delhi' as PhD institution, not a mark"
    ms_wb.close()

    pub_wb = _openpyxl.load_workbook(publications_path)
    assert "Amit Kumar Sharma" in pub_wb.sheetnames, "Expected one publications sheet per candidate"
    amit_pub_ws = pub_wb["Amit Kumar Sharma"]
    assert amit_pub_ws["B1"].value == "Amit Kumar Sharma", "Faculty name header missing on publications sheet"
    assert amit_pub_ws.cell(row=6, column=2).value == "Deep Learning for Resume Analysis", "Expected first real publication row to be populated"
    pub_wb.close()

    os.remove(marking_scheme_path)
    os.remove(publications_path)

    # Restore original function
    parsers.info_extractor.extract_candidate_profile_via_llm = original_llm_extractor
    
    print("\n-----------------------------------------------------")
    print("VERIFICATION COMPLETED SUCCESSFULLY!")
    print("All assertions passed. Overlap resolved and hybrid timeline engine calculations validated!")
    print("-----------------------------------------------------\n")

if __name__ == "__main__":
    run_verification()


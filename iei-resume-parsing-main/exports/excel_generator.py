import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Design Palette: Navy and Ice Blue
HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ZEBRA_FILL = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F497D")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F497D")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
REGULAR_FONT = Font(name="Calibri", size=11)

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

DOUBLE_BOTTOM_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='double', color='1F497D')
)

# ====================================================
# OFFICIAL FORM STYLE (matches the client's existing monochrome interview sheet)
# ====================================================
FORM_FONT_NAME = "Times New Roman"
FORM_BLACK_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
FORM_TITLE_FONT = Font(name=FORM_FONT_NAME, size=14, bold=True)
FORM_LABEL_FONT = Font(name=FORM_FONT_NAME, size=11, bold=True)
FORM_HEADER_FONT = Font(name=FORM_FONT_NAME, size=10, bold=True)
FORM_REGULAR_FONT = Font(name=FORM_FONT_NAME, size=10)
FORM_NOTE_FONT = Font(name=FORM_FONT_NAME, size=9, italic=True, color="595959")
FORM_UNVERIFIED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORM_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
FORM_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _sanitize_sheet_title(name, fallback="Sheet"):
    """Sheet titles must be <=31 chars and free of \\/?*[]: characters."""
    cleaned = re.sub(r"[\\*?:/\[\]]", "", name or fallback)[:28].strip()
    return cleaned if cleaned else fallback


def _guess_branch(raw_text):
    """
    Best-effort guess at engineering branch (COMP/MECH/EXTC/IT) from resume text.
    This is a guess, not a confirmed fact - the marking sheet shows it as a
    suggestion the panel should confirm, never silently presented as certain.
    """
    text_lower = (raw_text or "").lower()
    branch_keywords = [
        ("COMP", [r"\bcomputer\s+(?:science|engineering)\b", r"\bc\.?s\.?e\.?\b"]),
        ("MECH", [r"\bmechanical\s+engineering\b"]),
        ("EXTC", [r"\belectronics?\s+and\s+telecommunication\b", r"\be\s*&\s*tc\b", r"\bextc\b"]),
        ("IT", [r"\binformation\s+technology\b"]),
    ]
    for branch, patterns in branch_keywords:
        if any(re.search(p, text_lower) for p in patterns):
            return branch
    return None


def build_marking_scheme_excel(candidates_list, output_path, subject="", post="", interview_date=""):
    """
    Replicates the client's "Marking Scheme at the time of USSC Interview" sheet,
    but instead of pre-filling every candidate's qualification/experience columns
    with the maximum possible marks (which is what the original template does -
    every candidate row blindly mirrors the "Maximum Marks" row), this fills those
    columns with the ACTUAL data extracted from each candidate's resume, so the
    panel can see real facts and assign marks themselves rather than rubber-stamping
    a maximum score. Personal Interview and Total columns are left blank - those
    can only be filled in live, during the actual interview.

    NOTE on "FDP/STTP Attended" column: this replaces the original sheet's
    "Extra-Curricular Activities" column. Best-guess interpretation of a garbled
    client request - confirm the exact label with the client if it's wrong, it's
    a one-cell rename.
    """
    from database.db_manager import get_candidate_details

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "USSC Marking Scheme"
    ws.views.sheetView[0].showGridLines = False

    NUM_COLS = 13  # Sr.No, Name, Grad, PG, PhD, Industry Exp, Teaching Exp, Research Guide, Publications, FDP/STTP, Subject Matter, Comm Skills, Total

    # --- Title ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    title_cell = ws.cell(row=1, column=1, value="Marking Scheme at the time of USSC Interview")
    title_cell.font = FORM_TITLE_FONT
    title_cell.alignment = FORM_CENTER

    # --- Subject / Post / Date meta row ---
    meta_row = 3
    ws.cell(row=meta_row, column=1, value="Subject").font = FORM_LABEL_FONT
    ws.merge_cells(start_row=meta_row, start_column=2, end_row=meta_row, end_column=4)
    ws.cell(row=meta_row, column=2, value=subject or "Not specified").font = FORM_REGULAR_FONT

    ws.cell(row=meta_row, column=5, value="Post").font = FORM_LABEL_FONT
    ws.merge_cells(start_row=meta_row, start_column=6, end_row=meta_row, end_column=9)
    ws.cell(row=meta_row, column=6, value=post or "Not specified").font = FORM_REGULAR_FONT

    ws.cell(row=meta_row, column=10, value="Date").font = FORM_LABEL_FONT
    ws.merge_cells(start_row=meta_row, start_column=11, end_row=meta_row, end_column=NUM_COLS)
    ws.cell(row=meta_row, column=11, value=interview_date or "Not specified").font = FORM_REGULAR_FONT

    # --- Two-row grouped header ---
    header_row_1 = 5
    header_row_2 = 6

    def merged_group_header(start_col, end_col, text):
        ws.merge_cells(start_row=header_row_1, start_column=start_col, end_row=header_row_1, end_column=end_col)
        c = ws.cell(row=header_row_1, column=start_col, value=text)
        c.font = FORM_HEADER_FONT
        c.alignment = FORM_CENTER

    def vertical_merge_header(col, text):
        ws.merge_cells(start_row=header_row_1, start_column=col, end_row=header_row_2, end_column=col)
        c = ws.cell(row=header_row_1, column=col, value=text)
        c.font = FORM_HEADER_FONT
        c.alignment = FORM_CENTER

    vertical_merge_header(1, "Sr. No.")
    vertical_merge_header(2, "Name of the Candidate")
    merged_group_header(3, 5, "Qualification (Extracted from Resume)")
    merged_group_header(6, 10, "Teaching / Research / Administrative Experience (Extracted from Resume)")
    merged_group_header(11, 12, "Personal Interview (To Be Completed by Panel)")
    vertical_merge_header(13, "Total")

    sub_headers = {
        3: "Graduate\n(Institution)",
        4: "Post Graduate\n(Institution)",
        5: "Ph.D.\n(Institution)",
        6: "Industry\nExperience (Yrs)",
        7: "Teaching\nExperience (Yrs)",
        8: "Research Guide\n(PhD / PG / Projects)",
        9: "Standard Academic\nPublications (Count)",
        10: "FDP / STTP\nAttended (Count)",
        11: "Subject Matter\nof Interview",
        12: "Communication Skills &\nOverall Impression",
    }
    for col, text in sub_headers.items():
        c = ws.cell(row=header_row_2, column=col, value=text)
        c.font = FORM_HEADER_FONT
        c.alignment = FORM_CENTER

    for row in (header_row_1, header_row_2):
        for col in range(1, NUM_COLS + 1):
            ws.cell(row=row, column=col).border = FORM_BLACK_BORDER

    # --- Candidate data rows ---
    data_start_row = header_row_2 + 1
    for idx, candidate in enumerate(candidates_list):
        r = data_start_row + idx
        details = get_candidate_details(candidate["id"])
        ext_info = details.get("extracted_info", {}) if details else {}

        education = ext_info.get("education", {})
        guidance = ext_info.get("research_guidance", {})
        fdp_sttp = ext_info.get("fdp_sttp", {})

        row_values = [
            idx + 1,
            details.get("name", candidate.get("name", "Unknown")) if details else candidate.get("name", "Unknown"),
            education.get("ug_institution", "Not specified"),
            education.get("pg_institution", "Not specified"),
            education.get("phd_institution", "Not specified"),
            f"{ext_info.get('industry_experience_years', 0):.0f}",
            f"{ext_info.get('academic_experience_years', candidate.get('teaching_experience_years', 0)):.0f}",
            f"{guidance.get('phd_scholars_guided', 0)} PhD / {guidance.get('pg_students_guided', 0)} PG / {guidance.get('research_projects_supervised', 0)} Proj",
            f"{ext_info.get('publications', {}).get('total_publications', 0)}",
            f"{fdp_sttp.get('fdp_sttp_count', 0)}",
            "",  # Subject Matter of Interview - filled live during interview
            "",  # Communication Skills & Overall Impression - filled live during interview
            "",  # Total - depends on interview marks, filled live
        ]

        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = FORM_REGULAR_FONT
            cell.border = FORM_BLACK_BORDER
            cell.alignment = FORM_CENTER if col_idx != 2 else FORM_LEFT

    # --- Footnote ---
    footnote_row = data_start_row + len(candidates_list) + 2
    ws.merge_cells(start_row=footnote_row, start_column=1, end_row=footnote_row, end_column=NUM_COLS)
    note_cell = ws.cell(
        row=footnote_row, column=1,
        value=("Note: Qualification, Experience, and Publication columns are auto-extracted from each candidate's "
               "resume for the panel's reference. \"FDP / STTP Attended\" replaces the original sheet's "
               "\"Extra-Curricular Activities\" column - confirm this label with the client. "
               "Personal Interview and Total columns are intentionally left blank for the panel to complete live.")
    )
    note_cell.font = FORM_NOTE_FONT
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[footnote_row].height = 30

    # --- Column widths ---
    col_widths = {1: 7, 2: 24, 3: 16, 4: 16, 5: 16, 6: 11, 7: 11, 8: 16, 9: 12, 10: 12, 11: 16, 12: 18, 13: 9}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[header_row_1].height = 28
    ws.row_dimensions[header_row_2].height = 36

    wb.save(output_path)
    wb.close()
    return True


def build_publications_workbook(candidates_list, output_path):
    """
    Replicates the client's per-faculty publications detail sheet (Title of the
    Paper / Journal Name / Published Under / Year / Impact Factor / SCOPUS INDEX),
    one tab per candidate, auto-filled from each candidate's structured publication
    records. Rows flagged "Unverified" by the grounding check (i.e. the LLM claimed
    a paper that doesn't actually appear in the resume text) are highlighted, not
    hidden, so the panel knows to double check rather than trust it silently.
    """
    from database.db_manager import get_candidate_details

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # we only want one sheet per candidate, no blank default sheet

    NUM_COLS = 6  # Title, Journal Name, Published Under, Year, Impact Factor, SCOPUS INDEX
    MIN_BLANK_ROWS = 10  # keep the same "ready-to-fill" blank template feel as the client's original sheet

    for candidate in candidates_list:
        details = get_candidate_details(candidate["id"])
        if not details:
            continue
        ext_info = details.get("extracted_info", {})

        sheet_title = _sanitize_sheet_title(details.get("name", "Candidate"))
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = False

        # Header info block
        ws.cell(row=1, column=1, value="Name of The Faculty").font = FORM_LABEL_FONT
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=NUM_COLS + 1)
        ws.cell(row=1, column=2, value=details.get("name", "Unknown")).font = FORM_REGULAR_FONT

        # Designation: best-effort from the most recent extracted position, if any
        positions = ext_info.get("positions", [])
        designation = positions[-1]["designation"] if positions else "Not specified"
        ws.cell(row=2, column=1, value="Designation").font = FORM_LABEL_FONT
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=NUM_COLS + 1)
        ws.cell(row=2, column=2, value=designation).font = FORM_REGULAR_FONT

        # Branch: best-effort guess, clearly marked as needing confirmation
        guessed_branch = _guess_branch(details.get("raw_text", ""))
        branch_value = f"{guessed_branch} (auto-guessed - please confirm)" if guessed_branch else "Not specified - please confirm"
        ws.cell(row=3, column=1, value="Branch").font = FORM_LABEL_FONT
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=NUM_COLS + 1)
        ws.cell(row=3, column=2, value=branch_value).font = FORM_REGULAR_FONT

        # Table header
        header_row = 5
        headers = ["Title of the Paper", "Journal Name", "Published Under / Journal Details",
                   "Year of Publication", "Impact Factor (if Any)", "SCOPUS INDEX"]
        ws.cell(row=header_row, column=1, value="Sr.\nNo.").font = FORM_HEADER_FONT
        ws.cell(row=header_row, column=1).alignment = FORM_CENTER
        ws.cell(row=header_row, column=1).border = FORM_BLACK_BORDER
        for col_idx, h in enumerate(headers, 2):
            c = ws.cell(row=header_row, column=col_idx, value=h)
            c.font = FORM_HEADER_FONT
            c.alignment = FORM_CENTER
            c.border = FORM_BLACK_BORDER
        ws.row_dimensions[header_row].height = 30

        publications_detail = ext_info.get("publications_detail", [])
        num_rows = max(MIN_BLANK_ROWS, len(publications_detail))

        for i in range(num_rows):
            r = header_row + 1 + i
            record = publications_detail[i] if i < len(publications_detail) else None

            sr_cell = ws.cell(row=r, column=1, value=i + 1)
            sr_cell.font = FORM_REGULAR_FONT
            sr_cell.alignment = FORM_CENTER
            sr_cell.border = FORM_BLACK_BORDER

            if record:
                row_values = [
                    record.get("title", ""),
                    record.get("journal_name", ""),
                    record.get("published_under", ""),
                    record.get("year", ""),
                    record.get("impact_factor", ""),
                    record.get("scopus_indexed", ""),
                ]
            else:
                row_values = ["", "", "", "", "", ""]

            row_fill = FORM_UNVERIFIED_FILL if (record and not record.get("verified", True)) else None

            for col_idx, val in enumerate(row_values, 2):
                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.font = FORM_REGULAR_FONT
                cell.border = FORM_BLACK_BORDER
                cell.alignment = FORM_LEFT if col_idx in (2, 3, 4) else FORM_CENTER
                if row_fill:
                    cell.fill = row_fill
            if row_fill:
                sr_cell.fill = row_fill

        # Footnote on rows highlighted as unverified, only if any exist
        if any(not r.get("verified", True) for r in publications_detail):
            note_row = header_row + 1 + num_rows + 1
            ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=NUM_COLS + 1)
            note_cell = ws.cell(
                row=note_row, column=1,
                value="Highlighted rows could not be verified against the resume text - please double check these against the original document before relying on them."
            )
            note_cell.font = FORM_NOTE_FONT
            note_cell.alignment = Alignment(horizontal="left", wrap_text=True)
            ws.row_dimensions[note_row].height = 30

        # Column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 24
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12

    if not wb.sheetnames:
        # No candidates had retrievable details - still produce a usable empty sheet
        wb.create_sheet(title="No Candidates")

    wb.save(output_path)
    wb.close()
    return True

def build_evaluation_excel(candidates_list, output_path):
    """
    Generates a professionally styled Excel workbook containing:
    1. Master Ranking Sheet
    2. Individual evaluation scorecards for each candidate on separate tabs.
    """
    wb = openpyxl.Workbook()
    
    # 1. Create Master Ranking Sheet
    ws_master = wb.active
    ws_master.title = "Master Ranking"
    
    # Enable Gridlines
    ws_master.views.sheetView[0].showGridLines = True
    
    # Add Title
    ws_master["A1"] = "Faculty Recruitment Evaluation Dashboard"
    ws_master["A1"].font = TITLE_FONT
    ws_master["A2"] = "Master Candidate Ranking Sheet (Generated Offline)"
    ws_master["A2"].font = SUBTITLE_FONT
    
    # Table Headers
    headers = [
        "Rank", "Candidate Name", "Graduate", "Post Graduate", "PhD", 
        "Teaching Experience", "Research Guidance", "Publications", "Patents", 
        "Extra Curricular", "Total Score"
    ]
    
    header_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws_master.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        
    # Populate Data
    for row_idx, candidate in enumerate(candidates_list, 1):
        r_num = header_row + row_idx
        
        # Zebra striping
        row_fill = ZEBRA_FILL if row_idx % 2 == 0 else WHITE_FILL
        
        # Extract metadata
        from database.db_manager import get_candidate_details
        cand_detail = get_candidate_details(candidate["id"])
        ext_info = cand_detail.get("extracted_info", {}) if cand_detail else {}
        
        # Check degree indicators
        highest_q = candidate.get("highest_qualification", "None")
        grad_status = "Yes" if highest_q in ["Graduate", "Post Graduate", "PhD"] else "No"
        pg_status = "Yes" if highest_q in ["Post Graduate", "PhD"] else "No"
        phd_status = "Yes" if highest_q == "PhD" else "No"
        
        # Compile row values
        row_values = [
            candidate.get("rank", row_idx),
            candidate.get("name", "Unknown"),
            grad_status,
            pg_status,
            phd_status,
            f"{candidate.get('teaching_experience_years', 0.0)} yrs",
            f"{ext_info.get('research_guidance', {}).get('phd_scholars_guided', 0)} PhD / {ext_info.get('research_guidance', {}).get('pg_students_guided', 0)} PG",
            f"{ext_info.get('publications', {}).get('total_publications', 0)} papers",
            f"{ext_info.get('patents', {}).get('granted_patents', 0)} G / {ext_info.get('patents', {}).get('filed_patents', 0)} F",
            ", ".join(ext_info.get("extra_curricular_activities", [])) if ext_info.get("extra_curricular_activities") else "None",
            candidate.get("total_score", 0.0)
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws_master.cell(row=r_num, column=col_idx, value=val)
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            cell.fill = row_fill
            
            # Alignments
            if col_idx in [1, 3, 4, 5]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in [6, 7, 8, 9, 11]:
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
                
            # Bold for Total Score and Name
            if col_idx in [2, 11]:
                cell.font = BOLD_FONT

    # Adjust Master Column Widths
    for col in ws_master.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 3: # Skip titles
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_master.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Set Row height for headers
    ws_master.row_dimensions[header_row].height = 25
    
    # 2. Add individual tabs for each candidate
    for cand in candidates_list:
        # Load complete profile
        from database.db_manager import get_candidate_details
        details = get_candidate_details(cand["id"])
        if not details:
            continue
            
        ext_info = details.get("extracted_info", {})
        
        # Sheet title must be <= 31 chars and cannot contain special chars
        sheet_title = re.sub(r"[\\*?:/\[\]]", "", details["name"])[:28]
        ws_cand = wb.create_sheet(title=sheet_title)
        
        # Enable gridlines
        ws_cand.views.sheetView[0].showGridLines = True
        
        # Title block
        ws_cand["A1"] = f"Candidate Evaluation Scorecard"
        ws_cand["A1"].font = TITLE_FONT
        ws_cand["A2"] = f"Candidate Name: {details['name']}"
        ws_cand["A2"].font = BOLD_FONT
        
        # Meta info
        ws_cand["A4"] = "Email:"
        ws_cand["A4"].font = BOLD_FONT
        ws_cand["B4"] = details.get("email", "Not provided")
        ws_cand["B4"].font = REGULAR_FONT
        
        ws_cand["A5"] = "Phone:"
        ws_cand["A5"].font = BOLD_FONT
        ws_cand["B5"] = details.get("phone", "Not provided")
        ws_cand["B5"].font = REGULAR_FONT
        
        ws_cand["A6"] = "Highest Qual:"
        ws_cand["A6"].font = BOLD_FONT
        ws_cand["B6"] = details.get("highest_qualification", "None")
        ws_cand["B6"].font = REGULAR_FONT
        
        ws_cand["A7"] = "Teaching Experience:"
        ws_cand["A7"].font = BOLD_FONT
        ws_cand["B7"] = f"{details.get('teaching_experience_years', 0.0)} years"
        ws_cand["B7"].font = REGULAR_FONT
        
        # Score Summary Table Headers
        ws_cand["A9"] = "Evaluation Metric"
        ws_cand["A9"].font = HEADER_FONT
        ws_cand["A9"].fill = HEADER_FILL
        ws_cand["A9"].alignment = Alignment(horizontal="left")
        ws_cand["A9"].border = THIN_BORDER
        
        ws_cand["B9"] = "Calculated Score"
        ws_cand["B9"].font = HEADER_FONT
        ws_cand["B9"].fill = HEADER_FILL
        ws_cand["B9"].alignment = Alignment(horizontal="right")
        ws_cand["B9"].border = THIN_BORDER
        
        # Score rows
        scores_map = [
            ("Academic Qualification Score", details.get("qualification_score", 0.0)),
            ("Research Publications Score", details.get("publication_score", 0.0)),
            ("Teaching Experience Score", details.get("teaching_score", 0.0)),
            ("Research Guidance Score", details.get("research_guidance_score", 0.0)),
            ("Patents Score", details.get("patent_score", 0.0)),
            ("Extra-Curricular Score", details.get("extracurricular_score", 0.0)),
        ]
        
        for idx, (metric, score) in enumerate(scores_map, 10):
            ws_cand.cell(row=idx, column=1, value=metric).font = REGULAR_FONT
            ws_cand.cell(row=idx, column=1).border = THIN_BORDER
            
            c_score = ws_cand.cell(row=idx, column=2, value=score)
            c_score.font = REGULAR_FONT
            c_score.border = THIN_BORDER
            c_score.alignment = Alignment(horizontal="right")
            
        # Total Row
        total_row_idx = len(scores_map) + 10
        ws_cand.cell(row=total_row_idx, column=1, value="TOTAL SCORE").font = BOLD_FONT
        ws_cand.cell(row=total_row_idx, column=1).border = DOUBLE_BOTTOM_BORDER
        
        total_cell = ws_cand.cell(row=total_row_idx, column=2, value=details.get("total_score", 0.0))
        total_cell.font = BOLD_FONT
        total_cell.border = DOUBLE_BOTTOM_BORDER
        total_cell.alignment = Alignment(horizontal="right")
        
        # Section: Explainable AI Report
        report_start_row = total_row_idx + 3
        ws_cand.cell(row=report_start_row, column=1, value="Explainable AI Evaluation Report").font = SECTION_FONT
        
        # Print explanation report line-by-line
        report_text = details.get("explanation_report", "")
        report_lines = report_text.split("\n")
        
        for l_offset, line in enumerate(report_lines):
            r = report_start_row + 1 + l_offset
            # To look nice, we write line to cell A
            ws_cand.cell(row=r, column=1, value=line).font = Font(name="Courier New", size=10)
            # Merge cell A and B for long text lines
            ws_cand.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            
        # Set column widths
        ws_cand.column_dimensions["A"].width = 32
        ws_cand.column_dimensions["B"].width = 20
        ws_cand.column_dimensions["C"].width = 15
        ws_cand.column_dimensions["D"].width = 15
        
    # Save Workbook
    wb.save(output_path)
    wb.close()
    return True
